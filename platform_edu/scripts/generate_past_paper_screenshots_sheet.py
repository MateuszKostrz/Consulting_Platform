#!/usr/bin/env python3
"""Generate past-paper screenshot coverage spreadsheet from the database."""

import os
import sys
from pathlib import Path

import django
from django.db.models import Count, Q
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "platform_edu.settings")
django.setup()

from website.models import (  # noqa: E402
    Past_Paper_Videos,
    Past_Paper_Videos_AA_HL,
    Past_Paper_Videos_AA_SL,
    Past_Paper_Videos_AI_HL,
    Past_Paper_Videos_Biology_HL,
    Past_Paper_Videos_Biology_SL,
    Past_Paper_Videos_Chemistry_HL,
    Past_Paper_Videos_Chemistry_SL,
    Past_Paper_Videos_Comp_Sci_HL,
    Past_Paper_Videos_Comp_Sci_SL,
    Past_Paper_Videos_Physics_HL,
    Past_Paper_Videos_Physics_SL,
)

OUTPUT_PATH = PROJECT_ROOT / "data" / "past_paper_screenshots_coverage.xlsx"

HEADERS = [
    "Subject",
    "Session",
    "Past Paper Videos",
    "Markscheme",
]

SUBJECT_MODELS = [
    ("Math AI SL", Past_Paper_Videos),
    ("Math AI HL", Past_Paper_Videos_AI_HL),
    ("Math AA SL", Past_Paper_Videos_AA_SL),
    ("Math AA HL", Past_Paper_Videos_AA_HL),
    ("Physics SL", Past_Paper_Videos_Physics_SL),
    ("Physics HL", Past_Paper_Videos_Physics_HL),
    ("Comp Sci SL", Past_Paper_Videos_Comp_Sci_SL),
    ("Comp Sci HL", Past_Paper_Videos_Comp_Sci_HL),
    ("Biology SL", Past_Paper_Videos_Biology_SL),
    ("Biology HL", Past_Paper_Videos_Biology_HL),
    ("Chemistry SL", Past_Paper_Videos_Chemistry_SL),
    ("Chemistry HL", Past_Paper_Videos_Chemistry_HL),
]

IMAGEKIT_Q = Q(question_screenshots_url__icontains="imagekit")
IMAGEKIT_M = Q(markscheme_screenshots_url__icontains="imagekit")

STATUS_FILLS = {
    "Done": "C6EFCE",
    "Not done": "FFEB9C",
    "Partial": "FFC7CE",
}


def fmt_timezone(time_zone: str) -> str:
    if time_zone in ("null", "0", "", None):
        return ""
    return f"TZ{time_zone}"


def fmt_paper(paper: str) -> str:
    if paper in ("1", "2", "3"):
        return f"P{paper}"
    if paper in ("1A", "1B"):
        return f"P{paper}"
    if paper and paper.startswith("P"):
        return paper
    return f"P{paper}" if paper and paper != "null" else ""


def fmt_session(month: str, year: str, time_zone: str, paper: str) -> str:
    parts = [month, str(year)]
    tz = fmt_timezone(time_zone)
    if tz:
        parts.append(tz)
    paper_label = fmt_paper(paper)
    if paper_label:
        parts.append(paper_label)
    return " ".join(parts)


def status_label(done_count: int, total: int) -> str:
    if total == 0 or done_count == 0:
        return "Not done"
    if done_count == total:
        return "Done"
    return "Partial"


def sort_key(row: tuple) -> tuple:
    subject, session, _, _ = row
    parts = session.split()
    month = parts[0] if parts else ""
    year = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0
    session_order = {"May": 0, "November": 1, "Specimen": 2}.get(month, 3)
    tz_num = 0
    paper_order = 99
    for part in parts:
        if part.startswith("TZ") and part[2:].isdigit():
            tz_num = int(part[2:])
        if part.startswith("P"):
            paper_order = {"P1": 1, "P1A": 2, "P1B": 3, "P2": 4, "P3": 5}.get(part, 99)
    return (subject, -year, session_order, tz_num, paper_order)


def fetch_rows(subject: str, model) -> list[tuple]:
    if not hasattr(model, "question_screenshots_url"):
        return []

    rows = []
    papers = (
        model.objects.exclude(month="null")
        .values("month", "year", "time_zone", "paper")
        .annotate(
            total=Count("id"),
            question_done=Count("id", filter=IMAGEKIT_Q),
            markscheme_done=Count("id", filter=IMAGEKIT_M),
        )
        .order_by("year", "month", "time_zone", "paper")
    )
    for paper in papers:
        session = fmt_session(
            paper["month"],
            paper["year"],
            paper["time_zone"],
            paper["paper"],
        )
        rows.append(
            (
                subject,
                session,
                status_label(paper["question_done"], paper["total"]),
                status_label(paper["markscheme_done"], paper["total"]),
            )
        )
    return rows


def collect_rows() -> list[tuple]:
    rows: list[tuple] = []
    for subject, model in SUBJECT_MODELS:
        rows.extend(fetch_rows(subject, model))
    return sorted(rows, key=sort_key)


def build_workbook(rows: list[tuple]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Screenshot Coverage"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col_idx in (3, 4):
            status = row[col_idx - 1]
            if status in STATUS_FILLS:
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    "solid", fgColor=STATUS_FILLS[status]
                )

    column_widths = [16, 28, 18, 14]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"
    return wb


def main() -> None:
    rows = collect_rows()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(rows)
    wb.save(OUTPUT_PATH)
    done_q = sum(1 for r in rows if r[2] == "Done")
    done_m = sum(1 for r in rows if r[3] == "Done")
    print(f"Wrote {OUTPUT_PATH} ({len(rows)} rows)")
    print(f"  Past paper videos Done: {done_q}/{len(rows)}")
    print(f"  Markscheme Done: {done_m}/{len(rows)}")


if __name__ == "__main__":
    main()
