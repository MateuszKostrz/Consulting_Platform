#!/usr/bin/env python3
"""Generate EDUNADE content coverage tracking spreadsheet."""

import os
import sys
from pathlib import Path

import django
from django.db.models import Count
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "platform_edu.settings")
django.setup()

from website.models import (  # noqa: E402
    Past_Paper_Videos,
    Past_Paper_Videos_AA_HL,
    Past_Paper_Videos_AA_SL,
    Past_Paper_Videos_AI_HL,
    Past_Paper_Videos_Biology_HL,
    Past_Paper_Videos_Biology_SL,
    Past_Paper_Videos_Chemistry_HL,
    Past_Paper_Videos_Chemistry_SL,
    Past_Paper_Videos_Comp_Sci_HL,
    Past_Paper_Videos_Comp_Sci_SL,
    Past_Paper_Videos_Physics_HL,
    Past_Paper_Videos_Physics_SL,
)

OUTPUT_PATH = PROJECT_ROOT / "data" / "edunade_content_coverage.xlsx"

HEADERS = [
    "Subject",
    "Content Type",
    "Session",
    "Year",
    "Timezone",
    "Paper",
    "Status",
    "Notes",
]

SUBJECT_MODELS = [
    ("Math AI SL", Past_Paper_Videos),
    ("Math AI HL", Past_Paper_Videos_AI_HL),
    ("Math AA SL", Past_Paper_Videos_AA_SL),
    ("Math AA HL", Past_Paper_Videos_AA_HL),
    ("Physics SL", Past_Paper_Videos_Physics_SL),
    ("Physics HL", Past_Paper_Videos_Physics_HL),
    ("Comp Sci SL", Past_Paper_Videos_Comp_Sci_SL),
    ("Comp Sci HL", Past_Paper_Videos_Comp_Sci_HL),
    ("Biology SL", Past_Paper_Videos_Biology_SL),
    ("Biology HL", Past_Paper_Videos_Biology_HL),
    ("Chemistry SL", Past_Paper_Videos_Chemistry_SL),
    ("Chemistry HL", Past_Paper_Videos_Chemistry_HL),
]

# Manual markscheme tracking (not stored in DB).
MARKSCHEME_ROWS = [
    ("Math AI SL", "Markscheme", "May", 2025, "TZ1", "P1", "Complete", ""),
    ("Math AI SL", "Markscheme", "May", 2025, "TZ2", "P1", "Complete", ""),
    ("Math AI SL", "Markscheme", "May", 2025, "TZ3", "P1", "Complete", ""),
    ("Math AI SL", "Markscheme", "May", 2025, "TZ1", "P2", "Complete", ""),
    ("Math AI SL", "Markscheme", "May", 2025, "TZ2", "P2", "Complete", ""),
    ("Math AI SL", "Markscheme", "May", 2025, "TZ3", "P2", "Complete", ""),
    ("Math AI SL", "Markscheme", "November", 2025, "TZ1", "P1", "Complete", ""),
    ("Math AI SL", "Markscheme", "November", 2025, "TZ3", "P1", "Complete", ""),
    ("Math AI SL", "Markscheme", "November", 2025, "TZ1", "P2", "Complete", ""),
    ("Math AI SL", "Markscheme", "November", 2025, "TZ3", "P2", "Complete", ""),
    ("Math AI HL", "Markscheme", "May", 2025, "TZ1", "P1", "Complete", ""),
    ("Math AI HL", "Markscheme", "May", 2025, "TZ2", "P1", "Complete", ""),
    ("Math AI HL", "Markscheme", "May", 2025, "TZ3", "P1", "Pending", "Mati musi oddać opcje TZ3"),
    ("Math AI HL", "Markscheme", "May", 2025, "TZ1", "P2", "Complete", ""),
    ("Math AI HL", "Markscheme", "May", 2025, "TZ2", "P2", "Complete", ""),
    ("Math AI HL", "Markscheme", "May", 2025, "TZ3", "P2", "Pending", "Mati musi oddać opcje TZ3"),
    ("Math AI HL", "Markscheme", "May", 2025, "TZ1", "P3", "Complete", ""),
    ("Math AI HL", "Markscheme", "May", 2025, "TZ2", "P3", "Complete", ""),
    ("Math AI HL", "Markscheme", "May", 2025, "TZ3", "P3", "Pending", "Mati musi oddać opcje TZ3"),
    ("Math AI HL", "Markscheme", "November", 2025, "", "P1", "Complete", ""),
    ("Math AI HL", "Markscheme", "November", 2025, "", "P2", "Complete", ""),
    ("Math AI HL", "Markscheme", "November", 2025, "", "P3", "Complete", ""),
    ("Math AI HL", "Markscheme", "May", 2024, "TZ1", "P1", "Complete", ""),
    ("Math AA SL", "Markscheme", "May", 2025, "TZ1", "P1", "Complete", ""),
    ("Math AA SL", "Markscheme", "May", 2025, "TZ2", "P1", "Complete", ""),
    ("Math AA SL", "Markscheme", "May", 2025, "TZ3", "P1", "Pending", "Mati musi oddać opcje TZ3"),
    ("Math AA SL", "Markscheme", "May", 2025, "TZ1", "P2", "Error", "error"),
    ("Math AA SL", "Markscheme", "May", 2025, "TZ2", "P2", "Error", "error"),
    ("Math AA SL", "Markscheme", "May", 2025, "TZ3", "P2", "Pending", "Mati musi oddać opcje TZ3"),
    ("Math AA SL", "Markscheme", "November", 2025, "TZ1", "P1", "Complete", ""),
    ("Math AA SL", "Markscheme", "November", 2025, "TZ3", "P1", "Pending", "Mati musi oddać opcje TZ3"),
    ("Math AA SL", "Markscheme", "November", 2025, "TZ1", "P2", "Complete", ""),
    ("Math AA SL", "Markscheme", "November", 2025, "TZ3", "P2", "Pending", "Mati musi oddać opcje TZ3"),
    ("Math AA SL", "Markscheme", "May", 2024, "TZ1", "P1", "Complete", ""),
    ("Math AA SL", "Markscheme", "May", 2024, "TZ2", "P1", "Complete", ""),
    ("Math AA SL", "Markscheme", "May", 2024, "TZ1", "P2", "Complete", ""),
    ("Math AA SL", "Markscheme", "May", 2024, "TZ2", "P2", "Complete", ""),
    ("Math AA SL", "Markscheme", "November", 2024, "TZ1", "P1", "Complete", ""),
    ("Math AA SL", "Markscheme", "November", 2024, "TZ2", "P1", "Complete", ""),
    ("Math AA HL", "Markscheme", "May", 2025, "TZ1", "P1", "Complete", ""),
    ("Math AA HL", "Markscheme", "May", 2025, "TZ2", "P1", "Complete", ""),
    ("Math AA HL", "Markscheme", "May", 2025, "TZ3", "P1", "Pending", "Mati musi oddać opcje TZ3"),
    ("Math AA HL", "Markscheme", "May", 2025, "TZ1", "P2", "Complete", ""),
    ("Math AA HL", "Markscheme", "May", 2025, "TZ2", "P2", "Complete", ""),
    ("Math AA HL", "Markscheme", "May", 2025, "TZ3", "P2", "Pending", "Mati musi oddać opcje TZ3"),
    ("Math AA HL", "Markscheme", "May", 2025, "TZ1", "P3", "Complete", ""),
    ("Math AA HL", "Markscheme", "May", 2025, "TZ2", "P3", "Complete", ""),
    ("Math AA HL", "Markscheme", "May", 2025, "TZ3", "P3", "Pending", "Mati musi oddać opcje TZ3"),
    ("Math AA HL", "Markscheme", "November", 2025, "TZ1", "P1", "Complete", ""),
    ("Math AA HL", "Markscheme", "November", 2025, "TZ3", "P1", "Pending", "Mati musi oddać opcje TZ3"),
    ("Math AA HL", "Markscheme", "November", 2025, "TZ1", "P2", "Complete", ""),
    ("Math AA HL", "Markscheme", "November", 2025, "TZ3", "P2", "Pending", "Mati musi oddać opcje TZ3"),
    ("Math AA HL", "Markscheme", "November", 2025, "TZ1", "P3", "Complete", ""),
    ("Math AA HL", "Markscheme", "November", 2025, "TZ3", "P2", "Pending", "Mati musi oddać opcje TZ3"),
    ("Math AA HL", "Markscheme", "May", 2024, "TZ1", "P1", "Complete", ""),
    ("Physics SL", "Markscheme", "", "", "", "", "Pending", "No individual papers listed yet"),
    ("Physics HL", "Markscheme", "", "", "", "", "Pending", "No individual papers listed yet"),
    ("Comp Sci SL", "Markscheme", "May", 2025, "TZ2", "P1", "Complete", ""),
    ("Comp Sci SL", "Markscheme", "May", 2025, "TZ3", "P1", "Pending", "Mati musi oddać opcje TZ3"),
    ("Comp Sci SL", "Markscheme", "May", 2025, "TZ2", "P2", "Complete", ""),
    ("Comp Sci SL", "Markscheme", "May", 2025, "TZ3", "P2", "Pending", "Mati musi oddać opcje TZ3"),
    ("Comp Sci SL", "Markscheme", "November", 2025, "", "P1", "Complete", ""),
    ("Comp Sci SL", "Markscheme", "November", 2025, "", "P2", "Complete", ""),
    ("Comp Sci SL", "Markscheme", "May", 2024, "TZ1", "P1", "Complete", ""),
    ("Comp Sci SL", "Markscheme", "May", 2024, "TZ2", "P1", "Issue", "Pytania są wszystkie te same co w May 2024 TZ1 P1 — ten sam arkusz"),
    ("Comp Sci SL", "Markscheme", "May", 2024, "", "P2", "Pending", "Nie wiem jakie ma być TZ"),
    ("Comp Sci SL", "Markscheme", "November", 2024, "", "P1", "Pending", ""),
    ("Comp Sci SL", "Markscheme", "November", 2024, "", "P2", "Pending", ""),
    ("Chemistry SL", "Markscheme", "May", 2025, "TZ1", "P2", "Complete", ""),
    ("Chemistry SL", "Markscheme", "May", 2025, "TZ2", "P2", "Complete", ""),
    ("Chemistry SL", "Markscheme", "May", 2025, "TZ3", "P2", "Complete", ""),
]

SUBJECT_NOTES = {
    "Math AA HL": ("Subject", "Note", "nie ja robię"),
    "Comp Sci HL": ("Subject", "Pending", "No content listed yet"),
    "Biology SL": ("Subject", "Pending", "No content listed yet"),
    "Biology HL": ("Subject", "Pending", "No content listed yet"),
    "Chemistry SL": ("Subject", "Issue", "Nie ma opcji do wybrania Paper P1A"),
    "Chemistry HL": ("Subject", "Pending", "No content listed yet"),
}

# Past paper videos not yet in DB (Image Kit only).
MANUAL_VIDEO_ROWS = [
    ("Chemistry SL", "Past Paper Videos", "May", 2025, "TZ1", "P1A", "Issue", "Dodane na Image Kit, ale nie ma opcji do wybrania P1A — kontakt z Matim"),
    ("Chemistry SL", "Past Paper Videos", "May", 2025, "TZ2", "P1A", "Issue", "Dodane na Image Kit, ale nie ma opcji do wybrania P1A — kontakt z Matim"),
    ("Chemistry SL", "Past Paper Videos", "May", 2025, "TZ3", "P1A", "Issue", "Dodane na Image Kit, ale nie ma opcji do wybrania P1A — kontakt z Matim"),
    ("Chemistry SL", "Past Paper Videos", "May", 2025, "TZ1", "P1B", "Issue", "Dodane na Image Kit, ale nie ma opcji do wybrania P1B — kontakt z Matim"),
    ("Chemistry SL", "Past Paper Videos", "May", 2025, "TZ2", "P1B", "Issue", "Dodane na Image Kit, ale nie ma opcji do wybrania P1B — kontakt z Matim"),
    ("Chemistry SL", "Past Paper Videos", "May", 2025, "TZ3", "P1B", "Issue", "Dodane na Image Kit, ale nie ma opcji do wybrania P1B — kontakt z Matim"),
]

STATUS_FILLS = {
    "Complete": "C6EFCE",
    "Pending": "FFEB9C",
    "Error": "FFC7CE",
    "Issue": "FFC7CE",
    "Note": "DDEBF7",
}


def fmt_timezone(time_zone: str) -> str:
    if time_zone in ("null", "0", "", None):
        return ""
    return f"TZ{time_zone}"


def fmt_paper(paper: str) -> str:
    if paper in ("1", "2", "3"):
        return f"P{paper}"
    if paper in ("1A", "1B"):
        return f"P{paper}"
    if paper.startswith("P"):
        return paper
    return f"P{paper}"


def sort_key(row: tuple) -> tuple:
    _, _, session, year, timezone, paper, _, _ = row
    session_order = {"May": 0, "November": 1, "Specimen": 2}.get(session, 3)
    tz_num = int(timezone.replace("TZ", "")) if timezone and timezone.startswith("TZ") else 0
    paper_order = {"P1": 1, "P1A": 2, "P1B": 3, "P2": 4, "P3": 5}.get(paper, 99)
    return (year or 0, session_order, tz_num, paper_order)


def fetch_video_rows(subject: str, model) -> list[tuple]:
    rows = []
    papers = (
        model.objects.exclude(month="null")
        .values("month", "year", "time_zone", "paper")
        .annotate(question_count=Count("id"))
        .order_by("year", "month", "time_zone", "paper")
    )
    for paper in papers:
        rows.append(
            (
                subject,
                "Past Paper Videos",
                paper["month"],
                int(paper["year"]),
                fmt_timezone(paper["time_zone"]),
                fmt_paper(paper["paper"]),
                "Complete",
                f"{paper['question_count']} question videos on platform",
            )
        )
    return rows


def collect_rows() -> list[tuple]:
    markscheme_by_subject: dict[str, list[tuple]] = {}
    for row in MARKSCHEME_ROWS:
        markscheme_by_subject.setdefault(row[0], []).append(row)

    manual_videos_by_subject: dict[str, list[tuple]] = {}
    for row in MANUAL_VIDEO_ROWS:
        manual_videos_by_subject.setdefault(row[0], []).append(row)

    rows: list[tuple] = []
    for subject, model in SUBJECT_MODELS:
        if subject in SUBJECT_NOTES:
            content_type, status, note = SUBJECT_NOTES[subject]
            rows.append((subject, content_type, "", "", "", "", status, note))

        video_rows = fetch_video_rows(subject, model)
        if video_rows:
            rows.extend(sorted(video_rows, key=sort_key))
        elif subject not in SUBJECT_NOTES:
            rows.append((subject, "Past Paper Videos", "", "", "", "", "Pending", "No videos on platform yet"))

        manual_videos = manual_videos_by_subject.get(subject, [])
        if manual_videos:
            rows.extend(sorted(manual_videos, key=sort_key))

        markscheme_rows = markscheme_by_subject.get(subject, [])
        if markscheme_rows:
            rows.extend(sorted(markscheme_rows, key=sort_key))

    return rows


def build_workbook(rows: list[tuple]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Content Coverage"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        status = row[6]
        if status in STATUS_FILLS:
            fill = PatternFill("solid", fgColor=STATUS_FILLS[status])
            ws.cell(row=row_idx, column=7).fill = fill

    column_widths = [16, 18, 12, 8, 10, 8, 12, 60]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"
    return wb


def main() -> None:
    rows = collect_rows()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(rows)
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
